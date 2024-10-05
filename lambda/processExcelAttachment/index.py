import boto3
import os
import io
import base64
import email
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

s3 = boto3.client('s3')

def handler(event, context):
    print(f"Processing Excel attachment...")
    bucket_name = os.environ['BUCKET_NAME']
    message_id = event['messageId']
    attachment_filename = event['filename']
    
    pdf_key = f'invoices/{message_id}/{os.path.splitext(attachment_filename)[0]}.pdf'
    excel_data = None
    
    obj = s3.get_object(Bucket=bucket_name, Key=message_id)
    email_content = obj['Body'].read().decode('utf-8')
    for part in email.message_from_string(email_content).walk():
        if part.get_content_maintype() == 'application' and part.get_filename().endswith(('.xlsx', '.xls')):
            excel_data = part.get_payload(decode=True)
            break
    
    if excel_data:
        workbook = load_workbook(io.BytesIO(excel_data))
        sheet = workbook.active
        data = [list(row) for row in sheet.iter_rows(values_only=True)]
        
        buffer = io.BytesIO()
        pdf = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        table = Table(data)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        table.setStyle(style)
        pdf.build([table])
        
        pdf_data = buffer.getvalue()
        pdf_key = f'invoices/{message_id}.pdf'
        
        try:
            print(f"Saving PDF to bucket [{bucket_name}], at location [{pdf_key}]...")
            pdf_binary = pdf_data.encode('utf-8')
            
            s3.put_object(
                Bucket=bucket_name,
                Key=pdf_key,
                Body=pdf_binary,
                ContentType='application/pdf'
            )
            print(f"Successfully saved PDF to bucket [{bucket_name}], at location [{pdf_key}]!")
            result = {
                'statusCode': 200,
                'status': 'success',
                'pdfKey': pdf_key
            }
        except Exception as e:
            print(f"Error saving PDF: {str(e)}")
            result = {
                'statusCode': 400,
                'status': 'error',
                'error': str(e),
                'pdfKey': pdf_key
            }
        return result
    else:
        return {
            'statusCode': 404,
            'status': 'error',
            'body': f'No Excel attachment {attachment_filename} found'
        }