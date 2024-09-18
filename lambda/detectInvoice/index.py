import json
import boto3
import os
import email
from io import BytesIO
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

s3 = boto3.client('s3')

def convert_text_to_pdf(text):
    print("Converting text file to PDF...")
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = [Paragraph(text, styles['Normal'])]
    pdf.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
    
def convert_excel_to_pdf(excel_data):
    print("Converting excel file to PDF...")
    workbook = load_workbook(filename=BytesIO(excel_data))
    sheet = workbook.active
    
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=letter)
    
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append([str(cell) if cell is not None else '' for cell in row])
        
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements = []
    elements.append(table)
    pdf.build(elements)

    buffer.seek(0)
    return buffer.getvalue()

def handler(event, context):
    print("Executing detectInvoice: Subject does NOT contain 'UPDATE ACCOUNT ASSIGNMENTS'")
    
    bucket_name = os.environ['BUCKET_NAME']
    message_id = event['messageId']
    
    print(f"Retreiving email with messageId [{message_id}] from S3 bucket [{bucket_name}]")
    obj = s3.get_object(Bucket=bucket_name, Key=message_id)
    email_content = obj['Body'].read().decode('utf-8')
    
    print(f"Sucessfully retrieved email with messageId [{message_id}]! Parsing email...")
    msg = email.message_from_string(email_content)
    
    invoice_data = None
    invoice_type = None
    pdf_data = None
    
    print("Checking for attachments...")
    for part in msg.walk():
        if part.get_content_maintype() == 'application':
            filename = part.get_filename()
            if filename:
                if filename.endswith('.pdf'):
                    print("Found a PDF attachment!")
                    invoice_data = part.get_payload(decode=True)
                    invoice_type = invoice_data
                    pdf_data = invoice_data
                elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                    print("Found an Excel attachment!")
                    invoice_data = part.get_payload(decode=True)
                    invoice_type = 'excel'
                    pdf_data = convert_excel_to_pdf(invoice_data)
                    print("Converted excel to PDF successfully!")
    if not invoice_data:
        print("No attachment, checking email body...")
        if msg.is_multipart():
            for part in msg.get_payload():
                if part.get_content_type() == 'text/plain':
                    invoice_data = part.get_payload(decode=True)
                    invoice_type = 'text'
                    pdf_data = convert_text_to_pdf(invoice_data.decode('utf-8'))
                    print("Converted text to PDF sucessfully!")
    if pdf_data:
        pdf_key = f'invoices/{message_id}.pdf'
        print(f"Saving PDF Invoice at [{pdf_key}] to S3 bucket [{bucket_name}]")
        s3.put_object(Bucket=bucket_name, Key=pdf_key, Body=pdf_data)
        print(f"Saved invoice document sucessfully!")
        
        return {
            'statusCode': 200,
            'invoiceKey': pdf_key,
            'originalType': invoice_type
        }
    else:
        return {
            'statusCode': 400,
            'body': json.dumps('No invoice found in email')
        }
    